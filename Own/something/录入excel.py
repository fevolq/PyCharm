#!-*-coding:utf-8 -*-
#!python3.7
#!@Author:fuq666@qq.com
#!Updatetime:2020-08-28
#!Filename:将文件名录入至表格

import os,re
import openpyxl
import tkinter as tk
from tkinter import filedialog

#得到指定路径内的所有文件名
def get_name(file_path):
    l = []
    for folderName, subfolders, filenames in os.walk(file_path):  # 每次更改目录时要\\
        name = [folderName]         #文件夹名
        for filename in filenames:
            # print("包含的文件名", folderName, ":", filename)
            # filename = filename[:-4]
            name.append(filename)   #每个文件名
        l.append(name)
    pass
    return l

#将文件名处理，便于录入
def deal_name(l):
    dic_file = {}
    for i in l[1:]:
        file_name = i[0]
        file_name = os.path.basename(file_name)
        i.pop(0)
        name = i
        dic_file[file_name] = name
    return dic_file

def get_title(dic_file):
    #print('路径为：',dic_file)
    dic_file = re.sub(r'.*/','',dic_file)
    title = str(dic_file)
    return title

def search_code(name):
    r = re.compile(r'[a-z,A-Z]*\d*[a-z,A-Z]+-\d+')
    code = r.findall(name)
    if code:
        code = code[0]
        # print(code[0])
        return code
    else:
        # print('没有')
        return None
    pass

#保存至excel文件（每次运行都会重新录入所有，进行覆盖数据）
def save_excel(dic_file,excel_path,file_path):
    title = get_title(file_path)  # 用于获得表的取名
    #文件是否存在
    if os.path.exists(excel_path+'jav.xlsx'):
        pass
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '首页'
        wb.save(excel_path+'jav.xlsx')
    wb = openpyxl.load_workbook(excel_path+'jav.xlsx')
    try:
        sheet = wb[title]
        #首行是否存在
        if sheet['A1'].value == None:
            sheet['A1'] = '分类name'
            sheet['B1'] = '视频名称'
        else:
            pass
        #存入数据
        i = 2
        for name,videos in dic_file.items():
            for video in videos:
                sheet['A' + str(i)] = name
                sheet['B' + str(i)] = video
                i += 1
        wb.save(excel_path+'jav.xlsx')
    except:
        wb.create_sheet(title = title)
        wb.save(excel_path + 'jav.xlsx')
        save_excel(dic_file, excel_path,file_path)
        pass
    pass

#对excel文件进行一些处理
def deal_excel(excel_path,file_path):
    title = get_title(file_path)
    wb = openpyxl.load_workbook(excel_path + 'jav.xlsx')
    sheet = wb[title]
    #间距
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["D"].width = 10
    #A列字体
    # from openpyxl.styles import Font
    # italic24Font = Font(name='Calibri',size=24, bold=True)
    # sheet["A"].font = italic24Font
    #冻结第一行
    sheet.freeze_panes = 'A2'
    #设置C、D列
    sheet['C1'] = '来源'
    sheet['D1'] = '番号'
    i = 2
    while i <= sheet.max_row:
        name = sheet['B' + str(i)].value
        if 'HPJAV' in name:
            sheet['C' + str(i)] = 'JAV'
        elif 'Pornhub' in name:
            sheet['C' + str(i)] = 'Pornhub'
        else:
            sheet['C' + str(i)] = 'other'
        code = search_code(name)
        if code:
            sheet['D' + str(i)] = str(code)
        else:
            sheet['D' + str(i)] = '未找到'
        i += 1
    pass
    wb.save(excel_path+'jav.xlsx')

#选择文件夹
def select_path():
    root = tk.Tk()
    root.withdraw()
    Folderpath = filedialog.askdirectory()

    return Folderpath
    
def select_fold():
    print('请选择需要录入的文件夹')
    #a = input()
    fold = select_path()
    paths = [fold]
    while True:
        b = input('是否还需要录入其他文件夹（Y/N）：')
        if b == 'Y' or b == 'y':
            fold = select_path()
            paths.append(fold)
        else:
            break
    return paths

def main(excel_path,file_paths):
    for file_path in file_paths:
        l = get_name(file_path)
        dic_file = deal_name(l)
        save_excel(dic_file,excel_path,file_path)
        deal_excel(excel_path,file_path)
        # print(l)
    pass

if __name__ == '__main__':
    file_paths = select_fold()
    print('\n请选择需要存放excel的文件夹：')
    excel_path = select_path() + '/'
    main(excel_path,file_paths)
    print('文件保存在 {}'.format(excel_path))
    print('\nDone')
